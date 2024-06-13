import openpyxl as xl


wb = xl.load_workbook("transactions.xlsx")
# print(wb.sheetnames)

sheet = wb['Sheet1']


for row in range(2 , sheet.max_row+1) :
    cell_value = sheet.cell(row , 3).value
    corrected_value = (0.8*cell_value)
    sheet.cell(row,4).value = corrected_value
    
wb.save("transactions2.xlsx")